#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Copyright (C) 2025 sakito <sakito@sakito.com>
"""
Excelファイル処理

openpyxl利用
"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def create(data_dict):
    """
    作成
    """
    wb = Workbook()

    flag = True
    for sheet_name, data_lst in data_dict.items():

        if flag:
            ws = wb.active
            ws.title = sheet_name
            flag = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        for line in data_lst:
            ws.append(line)

    return wb


def save(wb, file_path):
    """
    保存
    """
    wb.save(file_path)


def get_color(color_num):
    """
    セルに色付与
    """
    fill = PatternFill(patternType='solid', fgColor=color_num)

    return fill
